''' Open an excel file table and save it as a json file

json structure based on column names
    - dot '.' indicates dictionary structure
    - name under braket "[...]" indicates semicolon ";" separated multi-value data
    - name starting with hash '#' is ignored
'''

from openpyxl import load_workbook

from vbrpytools.dicjsontools import merge_dict, save_json_file, create_nested_dict
from vbrpytools.misctools import get_args
from vbrpytools.misctools import force_stdout_encoding

class ExcelWorkbook():
    ''' Class to handle excel file
    '''
    def __init__(self, filename):
        ''' Open the excel file
        '''
        self._wb = load_workbook(filename = filename, data_only=True)

    @property
    def worksheets(self):
        ''' Get the list of worksheet names
        '''
        return self._wb.worksheets

    def table_ws(self, table_name):
        ''' Get the worksheet associated with a table name
        '''
        for ws in self.worksheets:
            if table_name in ws.tables.keys():
                return ws
        return None

    def get_table(self, table_name):
        ''' Get the table object
        '''
        ws = self.table_ws(table_name)
        if ws is None:
            return None
        return ws.tables[table_name]

    def dict_from_table(self, table_name,
                        nested = True, with_ignored = False):
        ''' Create a dictionary from an excel table
            column names are used as keys with the following rules:
                - dot '.' indicates dictionary structure - ignored if nested is False
                - name under braket "[...]" indicates semicolon ";" separated multi-value data
                - name starting with hash '#' is ignored - unless with_ignored is True
        Args:
            table_name (str): name of the table in the excel file
            nested (bool): if True, create nested dictionary based on column names
        Returns:
            list of dictionary: each entry corresponds to a row in the table
        '''

        ws = self.table_ws(table_name)
        in_table = ws.tables[table_name]
        in_range = ws[in_table.ref]
        column_names = in_table.column_names

        output = []
        for row in in_range[1:]: #skip first row which is the header
            output_entry = {}
            for name, row_cell in zip(column_names, row):
                cell_value = row_cell.value
                if (cell_value is not None) and (cell_value != '') and\
                   ((name[0] != '#') or with_ignored):
                    if name[0] == '[' and name[-1] == ']':
                        multivalue = True
                        name = name[1:-1]
                    else:
                        multivalue = False

                    keys = name.split('.') if nested else [name]

                    # Detect multi value content
                    if multivalue:
                        if str(cell_value).isnumeric():
                            cell_value = [cell_value]
                        else:
                            cell_value = str(cell_value).split(';')

                    output_entry_key = create_nested_dict(keys, cell_value)
                    output_entry = merge_dict(output_entry, output_entry_key)

            output.append(output_entry)
        return output

def _main():
    ''' Entry point of this module
    '''
    force_stdout_encoding()
    args_def = [(['-f', '--inputfile'     ], {'action':'store',                       'required':True , 'help':'Input Filename (xlsx) of the table',                           'metavar':'xxx.xlsx'}),
                (['-t', '--inputtable'    ], {'action':'store',                       'required':True , 'help':'Input table name',                                             'metavar':'xxx'     }),
                (['-o', '--outputfile'    ], {'action':'store',                       'required':True , 'help':'Output Filename (json)',                                       'metavar':'xxx.json'}),
                (['-p', '--preserve'      ], {'action':'store_true', 'default':False, 'required':False, 'help':'if set and output file exists, rename it by adding timestamp'                      }),
               ]
    args = get_args(args_def)

    wb = ExcelWorkbook(args.inputfile)
    output = wb.dict_from_table(args.inputtable)
    save_json_file(output, args.outputfile, preserve=args.preserve)

if __name__ == "__main__":
    _main()
